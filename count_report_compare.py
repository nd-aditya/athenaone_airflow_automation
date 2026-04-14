"""
Run on MAC after both count_report_mac.py and count_report_gcp.py have uploaded their CSVs to GCS.
Downloads both CSVs and produces a single comparison Excel with mismatch highlighting.

Usage:
    python count_report_compare.py
    python count_report_compare.py --out my_report.xlsx
"""
import argparse
import csv
import io
from datetime import datetime

from google.cloud import storage
from services.config import (
    COUNT_REPORT_GCS_BUCKET,
    COUNT_REPORT_MAC_GCS_PATH,
    COUNT_REPORT_GCP_GCS_PATH,
    DEIDENTIFIED_SCHEMA,
    GCP_REPORT_SCHEMAS,
)


def _download_csv(bucket_name: str, gcs_path: str) -> list[dict]:
    client = storage.Client()
    blob = client.bucket(bucket_name).blob(gcs_path)
    content = blob.download_as_text()
    return list(csv.DictReader(io.StringIO(content)))


def _int(val: str) -> int | None:
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def build_comparison(mac_rows: list[dict], gcp_rows: list[dict]) -> list[dict]:
    """
    Pivot GCP rows (one per table+schema_label) into wide format and join with MAC.
    Returns one row per table with mac_total, mac_active_y, gcp_<LABEL>_total, gcp_<LABEL>_active_y.
    """
    mac_by_table = {r["table"].upper(): r for r in mac_rows}

    # pivot GCP: {table_upper: {label: {total, active_y}}}
    gcp_pivot: dict[str, dict] = {}
    for r in gcp_rows:
        key = r["table"].upper()
        label = r["schema_label"]
        gcp_pivot.setdefault(key, {})[label] = {
            "total":    _int(r.get("total", "")),
            "active_y": _int(r.get("active_y", "")),
            "error":    r.get("error", ""),
        }

    gcp_labels = list(GCP_REPORT_SCHEMAS.keys())
    all_tables = sorted({t.upper() for t in list(mac_by_table.keys()) + list(gcp_pivot.keys())})

    rows = []
    for table in all_tables:
        mac = mac_by_table.get(table, {})
        row = {
            "table":        table,
            "mac_total":    _int(mac.get("total", "")),
            "mac_active_y": _int(mac.get("active_y", "")),
        }
        for label in gcp_labels:
            gcp = gcp_pivot.get(table, {}).get(label, {})
            row[f"gcp_{label}_total"]    = gcp.get("total")
            row[f"gcp_{label}_active_y"] = gcp.get("active_y")
        rows.append(row)

    return rows


def export_excel(rows: list[dict], path: str) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("Run: pip install openpyxl")

    gcp_labels = list(GCP_REPORT_SCHEMAS.keys())
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Count Comparison"

    # Styles
    white_bold   = Font(bold=True, color="FFFFFF")
    mac_fill     = PatternFill("solid", fgColor="1F4E79")   # dark blue
    gcp_fill     = PatternFill("solid", fgColor="375623")   # dark green
    mismatch_fill= PatternFill("solid", fgColor="FFCCCC")   # light red — diff between MAC and GCP
    na_fill      = PatternFill("solid", fgColor="EEEEEE")   # grey — missing/error
    thin         = Side(style="thin", color="CCCCCC")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    center       = Alignment(horizontal="center", vertical="center")

    # Title
    total_cols = 1 + 2 + len(gcp_labels) * 2
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    ws["A1"].value     = f"Priority Table Count Comparison  —  {ts}"
    ws["A1"].font      = Font(bold=True, size=13)
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 24

    # Headers
    headers = [("Table", None)]
    headers += [(f"MAC Total\n({DEIDENTIFIED_SCHEMA})", mac_fill), ("MAC\nActive(Y)", mac_fill)]
    for label in gcp_labels:
        headers += [(f"GCP {label}\nTotal", gcp_fill), (f"GCP {label}\nActive(Y)", gcp_fill)]

    for ci, (text, fill) in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=text)
        c.font      = white_bold
        c.fill      = fill or PatternFill("solid", fgColor="404040")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = border
    ws.row_dimensions[2].height = 36

    # Data
    for ri, row in enumerate(rows, 3):
        ci = 1
        c = ws.cell(row=ri, column=ci, value=row["table"])
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = border
        ci += 1

        mac_total  = row.get("mac_total")
        mac_active = row.get("mac_active_y")

        def write(val, compare=None):
            nonlocal ci
            c = ws.cell(row=ri, column=ci)
            if val is None:
                c.value = "N/A"
                c.fill  = na_fill
            else:
                c.value          = val
                c.number_format  = "#,##0"
                if compare is not None and compare != val:
                    c.fill = mismatch_fill
            c.alignment = center
            c.border    = border
            ci += 1

        write(mac_total)
        write(mac_active)

        for label in gcp_labels:
            gcp_total  = row.get(f"gcp_{label}_total")
            gcp_active = row.get(f"gcp_{label}_active_y")
            write(gcp_total,  compare=mac_total)
            write(gcp_active, compare=mac_active)

    # Column widths
    ws.column_dimensions["A"].width = 34
    for ci in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 18

    ws.freeze_panes = "B3"

    # Legend
    lws = wb.create_sheet("Legend")
    lws["A1"].value = "Colour"
    lws["B1"].value = "Meaning"
    lws["A1"].font  = Font(bold=True)
    lws["B1"].font  = Font(bold=True)
    for i, (fill, meaning) in enumerate([
        (mismatch_fill, "GCP count differs from MAC count"),
        (na_fill,       "Table missing or error"),
        (mac_fill,      "MAC columns"),
        (gcp_fill,      "GCP columns"),
    ], 2):
        lws.cell(row=i, column=1).fill  = fill
        lws.cell(row=i, column=2).value = meaning
    lws.column_dimensions["A"].width = 6
    lws.column_dimensions["B"].width = 44

    wb.save(path)
    print(f"Excel saved → {path}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", default=f"count_comparison_{datetime.now().strftime('%Y%m%d')}.xlsx")
    args = parser.parse_args()

    print(f"Downloading gs://{COUNT_REPORT_GCS_BUCKET}/{COUNT_REPORT_MAC_GCS_PATH} …")
    mac_rows = _download_csv(COUNT_REPORT_GCS_BUCKET, COUNT_REPORT_MAC_GCS_PATH)

    print(f"Downloading gs://{COUNT_REPORT_GCS_BUCKET}/{COUNT_REPORT_GCP_GCS_PATH} …")
    gcp_rows = _download_csv(COUNT_REPORT_GCS_BUCKET, COUNT_REPORT_GCP_GCS_PATH)

    print("Building comparison…")
    rows = build_comparison(mac_rows, gcp_rows)
    export_excel(rows, args.out)


if __name__ == "__main__":
    main()
