"""
Priority-table count report — MAC vs GCP comparison.

TWO-MACHINE WORKFLOW (when GCP port 3306 isn't directly reachable from MAC):
  Step 1 — on MAC:
      python services/count_report_service.py --mac-only --save-json mac_counts.json

  Step 2 — on GCP machine (copy this script + a minimal config_gcp.py there):
      python count_report_service.py --gcp-only --save-json gcp_counts.json

  Step 3 — copy gcp_counts.json back to MAC:
      scp user@172.16.2.42:~/gcp_counts.json .

  Step 4 — merge into Excel on MAC:
      python services/count_report_service.py --merge mac_counts.json gcp_counts.json --excel comparison.xlsx

SINGLE-MACHINE WORKFLOW (when MAC can reach GCP port 3306 directly):
      python services/count_report_service.py --excel comparison.xlsx
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime

from sqlalchemy import create_engine, text

# ── Config ────────────────────────────────────────────────────────────────────
# When running on GCP machine, these will be overridden by env vars or a local config.
try:
    from services.config import (
        MYSQL_USER,
        MYSQL_PASSWORD,
        MYSQL_HOST,
        DEIDENTIFIED_SCHEMA,
        GCP_MYSQL_HOST,
        GCP_MYSQL_USER,
        GCP_MYSQL_PASSWORD,
        GCP_MYSQL_PORT,
        GCP_REPORT_SCHEMAS,
    )
except ImportError:
    # Running standalone on GCP machine — set these directly
    import os
    MYSQL_USER          = os.getenv("MYSQL_USER", "nd-root-mysql")
    MYSQL_PASSWORD      = os.getenv("MYSQL_PASSWORD", "kmsamd89undsd4")
    MYSQL_HOST          = os.getenv("MYSQL_HOST", "localhost")
    DEIDENTIFIED_SCHEMA = os.getenv("DEIDENTIFIED_SCHEMA", "deidentified_merged")
    GCP_MYSQL_HOST      = ""
    GCP_MYSQL_USER      = ""
    GCP_MYSQL_PASSWORD  = ""
    GCP_MYSQL_PORT      = 3306
    GCP_REPORT_SCHEMAS  = {
        "TNG":     "tng_athena_one",
        "DCND":    "dcnd",
        "TNCPA":   "tncpa",
        "Raleigh": "raleigh",
    }

# Priority tables — matches TEST_TABLE_NAMES in extraction_dag.py
PRIORITY_TABLES = [
    "ALLERGY",
    "APPOINTMENT",
    "APPOINTMENTELIGIBILITYINFO",
    "APPOINTMENTNOTE",
    "APPOINTMENTVIEW",
    "CHART",
    "CHARTQUESTIONNAIRE",
    "CHARTQUESTIONNAIREANSWER",
    "CLINICALENCOUNTER",
    "CLINICALENCOUNTERDATA",
    "CLINICALENCOUNTERDIAGNOSIS",
    "CLINICALENCOUNTERDXICD10",
    "CLINICALENCOUNTERPREPNOTE",
    "CLINICALORDERTYPE",
    "clinicalprescription",
    "CLINICALRESULT",
    "CLINICALRESULTOBSERVATION",
    "CLINICALSERVICE",
    "CLINICALSERVICEPROCEDURECODE",
    "CLINICALTEMPLATE",
    "document",
    "FDB_RMIID1",
    "FDB_RNDC14",
    "ICDCODEALL",
    "INSURANCEPACKAGE",
    "medication",
    "PATIENT",
    "PATIENTALLERGY",
    "PATIENTALLERGYREACTION",
    "PATIENTFAMILYHISTORY",
    "PATIENTINSURANCE",
    "patientmedication",
    "PATIENTPASTMEDICALHISTORY",
    "PATIENTSOCIALHISTORY",
    "PATIENTSURGERY",
    "PATIENTSURGICALHISTORY",
    "PROCEDURECODE",
    "PROCEDURECODEREFERENCE",
    "SNOMED",
    "SOCIALHXFORMRESPONSE",
    "SOCIALHXFORMRESPONSEANSWER",
    "SURGICALHISTORYPROCEDURE",
    "visit",
    "VITALATTRIBUTEREADING",
    "VITALSIGN",
    "PROVIDER",
    "PROVIDERGROUP",
    "PATIENTPROBLEM",
    "PATIENTSNOMEDPROBLEM",
    "PATIENTSNOMEDICD10",
    "PATIENTGPALHISTORY",
]


# ── DB helpers ────────────────────────────────────────────────────────────────

def _count_table(conn, schema: str, table: str) -> tuple[int, int]:
    """Return (total, active_y). Returns (-1, -1) if table is missing or inaccessible."""
    try:
        total = conn.execute(
            text(f"SELECT COUNT(*) FROM `{schema}`.`{table}`")
        ).scalar() or 0
        active = conn.execute(
            text(f"SELECT COUNT(*) FROM `{schema}`.`{table}` WHERE nd_active_flag = 'Y'")
        ).scalar() or 0
        return int(total), int(active)
    except Exception:
        return -1, -1


def fetch_mac_counts(tables: list[str]) -> list[dict]:
    """Connect to local MySQL and return counts for DEIDENTIFIED_SCHEMA."""
    engine = create_engine(
        f"mysql+pymysql://{MYSQL_USER}:{MYSQL_PASSWORD}@{MYSQL_HOST}/",
        pool_pre_ping=True,
    )
    rows = []
    with engine.connect() as conn:
        for table in tables:
            total, active = _count_table(conn, DEIDENTIFIED_SCHEMA, table)
            rows.append({"table": table, "mac_total": total, "mac_active_y": active})
    engine.dispose()
    return rows


def fetch_gcp_counts(tables: list[str]) -> list[dict]:
    """Connect to GCP MySQL and return counts for all provider schemas."""
    if not GCP_MYSQL_HOST:
        raise ValueError("GCP_MYSQL_HOST is not configured in services/config.py")
    engine = create_engine(
        f"mysql+pymysql://{GCP_MYSQL_USER}:{GCP_MYSQL_PASSWORD}"
        f"@{GCP_MYSQL_HOST}:{GCP_MYSQL_PORT}/",
        pool_pre_ping=True,
    )
    rows = []
    with engine.connect() as conn:
        for table in tables:
            row: dict = {"table": table}
            for label, schema in GCP_REPORT_SCHEMAS.items():
                total, active = _count_table(conn, schema, table)
                row[f"gcp_{label}_total"]    = total
                row[f"gcp_{label}_active_y"] = active
            rows.append(row)
    engine.dispose()
    return rows


# ── Merge logic ───────────────────────────────────────────────────────────────

def merge_results(mac_rows: list[dict], gcp_rows: list[dict]) -> list[dict]:
    """Join MAC and GCP rows by table name into a single list."""
    gcp_by_table = {r["table"]: r for r in gcp_rows}
    merged = []
    for mac_row in mac_rows:
        gcp_row = gcp_by_table.get(mac_row["table"], {})
        merged.append({**mac_row, **{k: v for k, v in gcp_row.items() if k != "table"}})
    return merged


# ── Excel export ──────────────────────────────────────────────────────────────

def export_excel(rows: list[dict], path: str) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    gcp_labels = list(GCP_REPORT_SCHEMAS.keys())
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Count Comparison"

    # ── Styles ──
    header_font     = Font(bold=True, color="FFFFFF")
    mac_fill        = PatternFill("solid", fgColor="1F4E79")   # dark blue — MAC
    gcp_fill        = PatternFill("solid", fgColor="375623")   # dark green — GCP
    mismatch_fill   = PatternFill("solid", fgColor="FFCCCC")   # light red — count differs
    na_fill         = PatternFill("solid", fgColor="F2F2F2")   # grey — N/A
    center          = Alignment(horizontal="center", vertical="center")
    left            = Alignment(horizontal="left",   vertical="center")
    thin_side       = Side(style="thin", color="CCCCCC")
    thin_border     = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # ── Title row ──
    ws.merge_cells(f"A1:{get_column_letter(1 + 2 + len(gcp_labels) * 2)}1")
    title_cell = ws["A1"]
    title_cell.value = f"Priority Table Count Comparison  —  {ts}"
    title_cell.font  = Font(bold=True, size=13)
    title_cell.alignment = center
    ws.row_dimensions[1].height = 24

    # ── Header row ──
    headers = [("Table", None)]
    headers += [(f"MAC Total\n({DEIDENTIFIED_SCHEMA})", mac_fill), ("MAC Active(Y)", mac_fill)]
    for label in gcp_labels:
        headers += [(f"GCP {label}\nTotal", gcp_fill), (f"GCP {label}\nActive(Y)", gcp_fill)]

    for col_idx, (header_text, fill) in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header_text)
        cell.font      = header_font
        cell.fill      = fill or PatternFill("solid", fgColor="404040")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border
    ws.row_dimensions[2].height = 36

    # ── Data rows ──
    for row_idx, row in enumerate(rows, start=3):
        col = 1

        # Table name
        cell = ws.cell(row=row_idx, column=col, value=row["table"])
        cell.alignment = left
        cell.border    = thin_border
        col += 1

        mac_total  = row.get("mac_total",    -1)
        mac_active = row.get("mac_active_y", -1)

        def write_count(val: int, compare_val: int | None = None):
            nonlocal col
            c = ws.cell(row=row_idx, column=col)
            if val == -1:
                c.value     = "N/A"
                c.fill      = na_fill
            else:
                c.value     = val
                c.number_format = "#,##0"
                # Highlight mismatch vs MAC total (only for GCP total columns)
                if compare_val is not None and compare_val != -1 and val != compare_val:
                    c.fill  = mismatch_fill
            c.alignment = center
            c.border    = thin_border
            col += 1

        write_count(mac_total)
        write_count(mac_active)

        for label in gcp_labels:
            gcp_total  = row.get(f"gcp_{label}_total",    -1)
            gcp_active = row.get(f"gcp_{label}_active_y", -1)
            # Highlight if GCP total != MAC total
            write_count(gcp_total,  compare_val=mac_total)
            write_count(gcp_active, compare_val=mac_active)

    # ── Column widths ──
    ws.column_dimensions["A"].width = 34
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    # ── Freeze header rows ──
    ws.freeze_panes = "A3"

    # ── Legend sheet ──
    legend_ws = wb.create_sheet("Legend")
    legend_ws["A1"] = "Colour"
    legend_ws["B1"] = "Meaning"
    legend_ws["A1"].font = Font(bold=True)
    legend_ws["B1"].font = Font(bold=True)
    legends = [
        (mismatch_fill, "GCP count differs from MAC count"),
        (na_fill,        "Table not found or no access"),
        (mac_fill,       "MAC columns"),
        (gcp_fill,       "GCP columns"),
    ]
    for i, (fill, meaning) in enumerate(legends, start=2):
        legend_ws.cell(row=i, column=1).fill  = fill
        legend_ws.cell(row=i, column=2).value = meaning
    legend_ws.column_dimensions["A"].width = 6
    legend_ws.column_dimensions["B"].width = 40

    wb.save(path)
    print(f"Excel saved → {path}")


# ── JSON intermediate ─────────────────────────────────────────────────────────

def save_json(rows: list[dict], path: str) -> None:
    with open(path, "w") as f:
        json.dump({"generated_at": datetime.now().isoformat(), "rows": rows}, f, indent=2)
    print(f"JSON saved → {path}")


def load_json(path: str) -> list[dict]:
    with open(path) as f:
        data = json.load(f)
    return data["rows"]


# ── CLI ───────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Priority-table count comparison — MAC vs GCP",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--mac-only",  action="store_true", help="Fetch MAC counts only")
    mode.add_argument("--gcp-only",  action="store_true", help="Fetch GCP counts only")
    mode.add_argument(
        "--merge", nargs=2, metavar=("MAC_JSON", "GCP_JSON"),
        help="Merge two previously saved JSON files into Excel (no DB connection needed)"
    )

    parser.add_argument("--save-json", metavar="FILE", help="Save fetched counts as JSON (for later --merge)")
    parser.add_argument("--excel",     metavar="FILE", help="Export final comparison to Excel (.xlsx)")

    args = parser.parse_args()

    # ── Mode: merge two JSONs ──
    if args.merge:
        mac_json_path, gcp_json_path = args.merge
        mac_rows = load_json(mac_json_path)
        gcp_rows = load_json(gcp_json_path)
        merged   = merge_results(mac_rows, gcp_rows)
        out_path = args.excel or f"count_comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        export_excel(merged, out_path)
        return

    # ── Mode: live fetch ──
    tables = PRIORITY_TABLES

    if args.mac_only:
        print("Fetching MAC counts…")
        rows = fetch_mac_counts(tables)
        if args.save_json:
            save_json(rows, args.save_json)
        if args.excel:
            export_excel(rows, args.excel)
        elif not args.save_json:
            _print_simple(rows)

    elif args.gcp_only:
        print("Fetching GCP counts…")
        try:
            rows = fetch_gcp_counts(tables)
        except ValueError as exc:
            print(f"Config error: {exc}", file=sys.stderr)
            sys.exit(1)
        if args.save_json:
            save_json(rows, args.save_json)
        if args.excel:
            export_excel(rows, args.excel)
        elif not args.save_json:
            _print_simple(rows)

    else:
        # Both from this machine
        print("Fetching MAC counts…")
        mac_rows = fetch_mac_counts(tables)
        print("Fetching GCP counts…")
        try:
            gcp_rows = fetch_gcp_counts(tables)
        except ValueError as exc:
            print(f"Config error: {exc}", file=sys.stderr)
            sys.exit(1)
        rows = merge_results(mac_rows, gcp_rows)
        out_path = args.excel or f"count_comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        export_excel(rows, out_path)


def _print_simple(rows: list[dict]) -> None:
    for row in rows:
        parts = [f"{k}={v}" for k, v in row.items()]
        print("  ".join(parts))


if __name__ == "__main__":
    main()
